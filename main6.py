from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
import os
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
driver = webdriver.Chrome()

def wait_for_element(locator, timeout=10):
    return WebDriverWait(driver, timeout).until(EC.presence_of_element_located(locator))

file_path = 'rulings_data.xlsx'
if os.path.exists(file_path):
    workbook = load_workbook(file_path)
    worksheet = workbook.active
else:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["Span Text", "Subtitle Text", "Content Text","pagenumber"])  # Add headers

driver.get('https://rulings.cbp.gov/search?term=ruling&collection=ALL&sortBy=RELEVANCE&pageSize=100&page=1497')
number = 0
pagenumbr=1497
try:
    wait = WebDriverWait(driver, 10)
    
    while True:
        # Wait for the elements to be present
        elements = wait.until(
            EC.presence_of_all_elements_located((By.CSS_SELECTOR, ".mat-cell.cdk-cell.cdk-column-category.mat-column-category.ng-star-inserted"))
        )
        
        for element in elements:
            try:
                a_tag = element.find_element(By.TAG_NAME, "a")
                href = a_tag.get_attribute("href")
                
                
                a_tag.click()
               
                span = wait.until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, "span.title.mat-title"))
                )
                span_text = span.text
                

                div_subtitle = wait.until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, "div.subtitle.ng-star-inserted"))
                )
                subtitle_text = div_subtitle.text
              
                div_content = wait.until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, "div.scrollable-content.content"))
                )
                content_text = div_content.text
                
                print(f"Span text: {len(content_text)},{len(span_text)},{len(subtitle_text)}")
                print(number)

                # content_text = content_text.replace("\n", " ").replace("\r", " ")
                if len(content_text) > 32767:  # Excel's cell character limit
                    content_text = content_text[:32767]
                worksheet.append([span_text, subtitle_text, content_text,f"{pagenumbr}"])
                workbook.save(file_path)
                number+=1
               
                
            except Exception as e:
                print(f"An error occurred while clicking an element or getting span/div text: {e}")
                time.sleep(2)
        
        # Check if the "Next page" button is present and enabled
        driver.execute_script("window.scrollTo(0, 0);")
        time.sleep(1) 
        next_page_button = wait.until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "button[aria-label='Next page']"))
        )
        
        if next_page_button.get_attribute("disabled") == "true":
            break
        
        next_page_button.click()
        pagenumbr+=1
        time.sleep(3)
        mat_table = wait_for_element((By.CSS_SELECTOR, '.mat-table.cdk-table'))
        driver.execute_script("arguments[0].scrollTop = 0;", mat_table)
        time.sleep(1) 

except Exception as e:
    print(f"An error1 occurred: {e}")
    
    number+=1
finally:
    workbook.save(file_path)
    driver.quit()
